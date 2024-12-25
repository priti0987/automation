import openpyxl
from openpyxl import load_workbook
import base64
import os
class XlsLibrary:
    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'


def open_workbook(path_to_file):
    workbook = load_workbook(path_to_file)
    return workbook

def file_exist(path_to_file):
    file_exists = os.path.exists(path_to_file)
    return (file_exists)

def delete_file(path_to_file):
    os.remove(path_to_file)


def get_cell_value(file, sheet_name, row, column):
    workbook = load_workbook(file)
    worksheet1 = workbook.get_sheet_by_name(sheet_name)
    cell_value1 = worksheet1.cell(int(row), int(column)).value
    return cell_value1


def write_data(file, sheetname, row, col, varValue):
    workbook = load_workbook(file)
    worksheet1 = workbook.get_sheet_by_name(sheetname)
    worksheet1.cell(int(row), int(col)).value = varValue
    workbook.save(file)

def save_new_excel_sheet(file,sheetname,row,col,value):
    workbook = load_workbook(file)
    worksheet=workbook.create_sheet(sheetname)
    worksheet.cell(int(row), int(col)).value = value
    workbook.save(file)

def save_excel( file):
    """
        Save the excel file after writing the data.
        Example:
        Update existing file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        Save in new file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  D:\\Test\\ExcelRobotNewFile.xlsx                   |
        """
    workbook = load_workbook(file)
    workbook.save(file)


def get_row_count(file, sheet_name):
    workbook = load_workbook(file)
    worksheet1 = workbook.get_sheet_by_name(sheet_name)
    return    worksheet1.max_row

def get_row_data(file, sheet_name):
    workbook = load_workbook(file)
    worksheet1 = workbook.get_sheet_by_name(sheet_name)
    return    worksheet1.rows

def remove_sheet(file,sheetname):
    workbook = load_workbook(file)
    activeSheets =  workbook.get_sheet_names()
    print(activeSheets)
    if sheetname in activeSheets:
        worksheet1 = workbook.get_sheet_by_name(sheetname)
        workbook.remove_sheet(worksheet1)
        workbook.save(file)
    else:
        print("No Sheet with this name.... ")


def encode_my_Data(data):

    sample_string = data
    sample_string_bytes = sample_string.encode("ascii")
    base64_bytes = base64.b64encode(sample_string_bytes)
    base64_string = base64_bytes.decode("ascii")
    return base64_string
